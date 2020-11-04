import openpyxl as xl
import os
import re
import time
from pathlib import Path
#处理.xlsx
'''
    sel.excel   wb对象
    self.sheet  ws对象
 '''
class Deal_Excel:
    def __init__(self, filename, sheet_name = None):
        path = Path(filename)
        #文件名
        self.filename = filename
        
        #excel对象
        self.excel = xl.Workbook() if not path.exists() else xl.load_workbook(filename)
            
        #支持的excel类型
        self.file_type_list = ["*.xlsm", "*.xlsx", "*.xltx", "*.xltm"]
  
        #sheet对象
        #self.excel[]=self.excel.get_sheet_by_name
        self.sheet = self.excel.create_sheet(index = 0 ,title = sheet_name) if sheet_name not in self.excel.sheetnames else self.excel[sheet_name]#[]=get_sheet_by_name
        
        print(f"max_column={self.sheet.max_column},max_row={self.sheet.max_row}")
        
    '''
    析构函数 貌似调用前就先结束了
    def __del__(self):
        print(f"self.filename={self.filename}")
        self.excel.save(self.filename)
    '''
    #获取某个文件目录下的excel
    def Get_All_Excel(self):
        #扫描当前目录下的.excel
        for file_type in self.file_type_list:
            for file in self.dir.glob(file_type): #会返回找到的对象,
                print(file)
        
     #附加数据
    def Append_Line(self, contain = None):
        #会在文件的最后一行(从没数据开始的第一行,添加数据)
        self.sheet.append(contain)

    #按照行读sheet
    def Read_Line(self, min=0, max=None):
        max = max or self.sheet.max_row
        #访问指定的sheet的内容,通过指定范围(行 → 行)
        for row in self.sheet.iter_rows(min_row=min, max_col=self.sheet.max_column, max_row=max):
            #print(cell)结果为<Cell Sheet1.A1>指定了元素   
            print([ cell.value for cell in row ])
       
    #读sheet的第col列
    def Read_Col(self, col , start_row = 2):
        #取值时sheet["a1"]=sheet.cell(1,1)
        print([self.sheet.cell(row,col).value for row in range(start_row,self.sheet.max_row+1)])
    
if __name__ == "__main__":
    filename = "genlist.xlsx"
    t_excel = Deal_Excel(filename=filename, sheet_name = "200-interface")
    #t_excel.Append_Line(contain = ["hello","kali","mimi"])
    #t_excel.Read_Line()
    #t_excel.Read_Col(col=2)
    method = "POST"
    url = "/api/v1/namespaces/namespace/interfaces"
    i = 1
    code = 200
    #f-string大括号外如果需要显示大括号，则应输入连续两个大括号 {{ 和 }}：
    for i in range(1,203):
        body=f'''{{"vlanId":{i},
        "name":"veth.{i}",
        "ifType":"VLANIF",
        "description":"",
        "reverseRouteEnable":false,
        "ipv4":{{"ipv4Mode":"STATIC","staticIp":[]}},
        "ipv6":{{"ipv6Mode":"STATIC","staticIp":[],"ipv6Param":{{"mtu":1500,"enable":true}}}},
        "mtu":1500,
        "manage":{{"https":true,"ping":true,"ssh":false}}
        }}'''
        #print([method,url,body,code])
        t_excel.Append_Line(contain = [method,url,body,code])
    t_excel.excel.save(filename)